import datetime
import openpyxl
import time
import os
import shutil
from selenium import webdriver
from selenium.webdriver.support.select import Select
import sys


#获取yyyymm和now
today = datetime.date.today()
yyyy_now = str(today.year)
mm_now = str(today.month)
yyyymm = input('请输入数据年月（格式yyyy-mm），当月请直接按回车键：')
if yyyymm == '':
    yyyy = str(today.year)
    mm = str(today.month)
    mm_2019 = (int(yyyy) - 2019) * 12 + int(mm)
    yyyymm = yyyy + '-' + mm
else :
    yyyy = yyyymm[0:4]
    mm = yyyymm[5:]
    mm_2019 = (int(yyyy) - 2019) * 12 + int(mm)



#读取信息表格中的内容
info_xlsx = openpyxl.load_workbook('信息表格.xlsx',data_only = True)
info_sht = info_xlsx['Sheet1']
PROSAS_path = info_sht.cell(1,4).value
path_read = info_sht.cell(2,4).value
path_schedul = info_sht.cell(3,4).value
path_sum = info_sht.cell(4,4).value
dir_read = os.path.join(path_read,yyyymm)
dir_save = os.path.join(path_read,yyyymm,'已填报')
while os.path.exists(dir_save) == False:
    os.makedirs(dir_save)
username = info_sht['B1'].value
password = info_sht['B2'].value
shop_name_id = {}
for i in range(4,info_sht.max_row+1):
  shop_name_id[info_sht.cell(i,1).value] = info_sht.cell(i,2).value
info_xlsx.save('信息表格.xlsx')


#打开进度表、读取sht
schedul_xlsx = openpyxl.load_workbook(path_schedul,data_only = True)
schedul_sht = schedul_xlsx['卫生监督进度表']


#读取月份文件夹下的所有文件
files_xlsx = []
files_save = []
for root,dirs,files in os.walk(dir_read,topdown = False):
    for file in files:
        files_xlsx.append(os.path.join(root,file))
while ('已填报' in files_xlsx[0]) == True:
    del files_xlsx[0]
    if files_xlsx == []:
        break


#判断符合或者不符合的函数，x是项的序号，y是符合或者不符合或者合理缺项:
def trueorfalse(x,y):
    if y == "符合":
        return typ_list[int(x/2-1)]
    if y == "合理缺项":
        return '99'
    if y == "不符合":
        return '0'


#填入结果的函数:
def result(score,i):
    if score!='0':
        el_score=browser.find_element_by_css_selector(name_score+value_score)
        browser.execute_script("arguments[0].scrollIntoView();",el_score)
        browser.execute_script("arguments[0].click();",el_score)
    else:
        el_score=browser.find_element_by_css_selector(name_score+value_score)
        browser.execute_script("arguments[0].scrollIntoView();",el_score)
        browser.execute_script("arguments[0].click();",el_score)
        el_explain=browser.find_element_by_css_selector(input_score)
        el_explain.send_keys(list_score[i+1])


if files_xlsx == []:
    print('未发现卫生监督记录，即将进行卫生监督统计...')
else:
    tian_or_not = input('是否启动PROSAS填报？（y/n）')
    if tian_or_not == 'n':
        print('即将进行卫生监督统计，但部分监督记录尚未填报，该记录将不会纳入统计，请注意！')
    elif tian_or_not != 'n' and tian_or_not != 'y':
        print('您的输入有误，程序终止，请重新启动程序并正确输入！')
        sys.exit()
    elif tian_or_not == 'y':
        '''#开浏览器、打开网页
        browser = webdriver.Firefox()
        browser.maximize_window()
        browser.get(PROSAS_path)

        #登录账号密码
        while 1:
            try:
                el_username=browser.find_element_by_id('username')
                el_username.send_keys(username)#输入用户名
                print('输入账号………………成功')
                el_password=browser.find_element_by_id('password')
                el_password.send_keys(password)#输入密码
                print('输入密码………………成功')
                submit=browser.find_element_by_name('submit')
                submit.click()#登录按钮
                print('登录……………………成功')
                break
            except:
                print('登录失败，请检查网络')

        #找到监督评分→点击
        el_ywjg=browser.find_element_by_id('heTab105')
        el_ywjg.click()#点击“卫生监督”按钮
        el_rcwsjd=browser.find_element_by_partial_link_text('日常卫生监督')
        el_rcwsjd.click()#点击日常卫生监督按钮
        el_jdpf=browser.find_element_by_partial_link_text('监督评分')
        el_jdpf.click()#点击监督评分按钮
'''
        for file_name in files_xlsx:
            #加载文件
            wb=openpyxl.load_workbook(file_name)
            sheet=wb['Sheet1']
            
            #判断卫生监督类型赋值给typ_jd和typ_list变量并确定各项的分值
            if '餐饮' in file_name:
                typ_jd='餐饮服务'
                typ_list=['※','※','※','2','5','※','2','2','2','2','※','※','※','2','2','1','5','1','2','5','5','2','※','5','10','5','5','1','2','5','2','5','2','※','5','2','※','2','5','2','2','5','5','5','2','※','2','2','2','2','2','1','2','5','2','5','2','2','2','2','2','2','2','※','2','2','5','5','※','2','5','5','5','2']
            elif '生产' in file_name:
                typ_jd='食品生产'
                typ_list=['※','※','※','2','5','※','5','5','2','2','※','※','※','※','5','5','5','2','※','※','5','5','2','2','5','2','2','5','10','5','※','※','5','10','10','5','※','5','10','※','5','5','5','2','2','※','2','2','2','5','10','5','※','※','5','5','5']
            elif '饮用水' in file_name:
                typ_jd='饮用水供应'
                typ_list=['※','※','※','5','10','5','5','5','5','5','※','※','5','5','5','※','5','2','5','5','2','5','5','10','10','5','※','2','※','5','5','※','2','2','10','2']
            elif '销售' in file_name:
                typ_jd='食品销售'
                typ_list=['※','※','※','5','5','5','5','5','5','※','※','10','10','5','5','5','※','5','5','2','5','2','2','※','5','5','10','5','5','5','5','10','5']
            elif '住宿' in file_name:
                typ_jd='住宿业'
                typ_list=['※','※','※','5','10','5','5','2','※','5','10','3','5','3','※','2','10','※','10','10','5','5','※','5','5','5','2','3','3','3','3','3','2','3','5']
            elif '候车（机、船）室' in file_name:
                typ_jd='候车（机、船）室'
                typ_list=['※','※','※','5','10','5','5','2','※','5','10','3','5','3','※','2','10','※','10','10','5','5','※','5','5','5','5','※','10']
            elif '文化娱乐场所' in file_name:
                typ_jd='文化娱乐场所'
                typ_list=['※','※','※','5','10','5','5','2','※','5','10','3','5','3','※','2','10','※','10','10','5','5','※','5','5','5','3','10']
            elif '美容美发场所' in file_name:
                typ_jd='美容美发场所'
                typ_list=['※','※','※','5','10','5','5','2','※','5','10','3','5','3','※','2','10','※','10','10','5','5','※','5','5','5','5','6','※']
            elif '沐浴场所' in file_name:
                typ_jd='沐浴场所'
                typ_list=['※','※','※','5','10','5','5','2','※','5','10','3','5','3','※','2','10','※','10','10','5','5','※','5','5','5','5','2','5','5','3']
            elif '游泳场所' in file_name:
                typ_jd='游泳场所'
                typ_list=['※','※','※','5','10','5','5','2','※','5','10','3','5','3','※','2','10','※','10','10','5','5','※','5','5','5','2','3','3']

            #将一次卫生监督结果存入list_score变量，然后将变量写入浏览器表单
            for r in range(2,sheet.max_row+1):
                list_score=[]
                for c in range(1,sheet.max_column+1):
                    list_score.append(sheet.cell(r,c).value)
                for i in range(len(list_score)):
                    if list_score[i] == '不符合' and list_score[i+1] == None:
                        list_score[i+1]=' '

                '''time.sleep(8)
                browser.switch_to.default_content()
                el_frame=browser.find_element_by_class_name('iframeClass')
                browser.switch_to.frame(el_frame)
                el_No=browser.find_element_by_name('cardNo')
                el_No.clear()
                el_No.send_keys(shop_name_id[list_score[1]])
                el_startDate=browser.find_element_by_name('startDate')
                sDate=datetime.datetime.now()-datetime.timedelta(days=365)#起始日期（当前时间往前推365天）
                browser.execute_script('arguments[0].removeAttribute(\"readonly\")',el_startDate)
                el_startDate.clear()
                el_startDate.send_keys(str(sDate.year)+'-'+str(sDate.month)+'-'+str(sDate.day))#输入起始日期（当前时间往前推365天）
                el_submit=browser.find_element_by_xpath("//input[@value='查询']")
                el_submit.click()
                time.sleep(5)
                el_add=browser.find_element_by_xpath("//i[@title='监督打分']")
                el_add.click()
                time.sleep(3)
                el_type=browser.find_element_by_id('itemCode')
                Select(el_type).select_by_visible_text(typ_jd)
                el_typechange=browser.find_element_by_xpath("//a[contains(text(),'修改监督评分表类型')]")
                el_typechange.click()
                time.sleep(0.3)
                browser.switch_to.default_content()
                el_frame_type=browser.find_element_by_css_selector("[src='/prosas/dailySup/listNoQuery.html?menuId=8B4C90F4861945B59DD330DA2378B103']")
                browser.switch_to.frame(el_frame_type)
                #el_typesubmit=browser.find_element_by_css_selector("button")
                el_typesubmit=browser.find_element_by_css_selector("button[class='aui_state_highlight'][type='button']")
                #el_typesubmit.click()
                browser.execute_script("$(arguments[0]).click()",el_typesubmit)

                for i in range(2,len(list_score),2):
                  i = int(i)
                  time.sleep(0.5)
                  score=trueorfalse(i,list_score[i])
                  if typ_jd=='餐饮服务':
                    if 0<i<9:
                      name_score="[name='score01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 9<i<23:
                      name_score="[name='score02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 23<i<31:
                      name_score="[name='score03"+(str(int((i-22)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input03"+(str(int((i-22)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 31<i<35:
                      name_score="[name='score04"+(str(int((i-30)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input04"+(str(int((i-30)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 35<i<45:
                      name_score="[name='score05"+(str(int((i-34)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input05"+(str(int((i-34)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 45<i<53:
                      name_score="[name='score06"+(str(int((i-44)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input06"+(str(int((i-44)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 53<i<63:
                      name_score="[name='score07"+(str(int((i-52)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input07"+(str(int((i-52)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 63<i<67:
                      name_score="[name='score08"+(str(int((i-62)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input08"+(str(int((i-62)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 67<i<73:
                      name_score="[name='score09"+(str(int((i-66)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input09"+(str(int((i-66)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 73<i<91:
                      name_score="[name='score10"+(str(int((i-72)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input10"+(str(int((i-72)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 91<i<111:
                      name_score="[name='score11"+(str(int((i-90)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input11"+(str(int((i-90)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 111<i<123:
                      name_score="[name='score12"+(str(int((i-110)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input12"+(str(int((i-110)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 123<i<137:
                      name_score="[name='score13"+(str(int((i-122)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input13"+(str(int((i-122)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 137<i<143:
                      name_score="[name='score14"+(str(int((i-136)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input14"+(str(int((i-136)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 143<i<147:
                      name_score="[name='score15"+(str(int((i-142)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input15"+(str(int((i-142)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 147<i<149:
                      name_score="[name='score16"+(str(int((i-146)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input16"+(str(int((i-146)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                  elif typ_jd=='食品生产':
                    if 0<i<9:
                      name_score="[name='score01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 9<i<21:
                      name_score="[name='score02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 21<i<23:
                      name_score="[name='score03"+(str(int((i-20)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input03"+(str(int((i-20)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 23<i<47:
                      name_score="[name='score04"+(str(int((i-22)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input04"+(str(int((i-22)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 47<i<57:
                      name_score="[name='score05"+(str(int((i-46)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input05"+(str(int((i-46)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 57<i<79:
                      name_score="[name='score06"+(str(int((i-56)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input06"+(str(int((i-56)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 79<i<85:
                      name_score="[name='score07"+(str(int((i-78)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input07"+(str(int((i-78)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 85<i<95:
                      name_score="[name='score08"+(str(int((i-84)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input08"+(str(int((i-84)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 95<i<101:
                      name_score="[name='score09"+(str(int((i-94)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input09"+(str(int((i-94)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 101<i<105:
                      name_score="[name='score10"+(str(int((i-100)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input10"+(str(int((i-100)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 105<i<115:
                      name_score="[name='score11"+(str(int((i-104)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input11"+(str(int((i-104)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                  elif typ_jd=='饮用水供应':
                    if 0<i<9:
                      name_score="[name='score01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 9<i<21:
                      name_score="[name='score02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 21<i<23:
                      name_score="[name='score03"+(str(int((i-20)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input03"+(str(int((i-20)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 23<i<25:
                      name_score="[name='score04"+(str(int((i-22)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input04"+(str(int((i-22)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 25<i<33:
                      name_score="[name='score05"+(str(int((i-24)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input05"+(str(int((i-24)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 33<i<37:
                      name_score="[name='score06"+(str(int((i-32)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input06"+(str(int((i-32)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 37<i<43:
                      name_score="[name='score07"+(str(int((i-36)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input07"+(str(int((i-36)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 43<i<63:
                      name_score="[name='score08"+(str(int((i-42)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input08"+(str(int((i-42)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 63<i<65:
                      name_score="[name='score09"+(str(int((i-62)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input09"+(str(int((i-62)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 65<i<75:
                      name_score="[name='score10"+(str(int((i-64)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input10"+(str(int((i-64)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                  elif typ_jd=='食品销售':
                    if 0<i<9:
                      name_score="[name='score01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 9<i<19:
                      name_score="[name='score02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 19<i<21:
                      name_score="[name='score03"+(str(int((i-18)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input03"+(str(int((i-18)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 21<i<39:
                      name_score="[name='score04"+(str(int((i-20)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input04"+(str(int((i-20)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 39<i<47:
                      name_score="[name='score05"+(str(int((i-38)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input05"+(str(int((i-38)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 47<i<67:
                      name_score="[name='score06"+(str(int((i-46)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input06"+(str(int((i-46)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                  elif typ_jd=='住宿业':
                    if 0<i<9:
                      name_score="[name='score01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input02"+(str(int(i/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 9<i<19:
                      name_score="[name='score02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 19<i<27:
                      name_score="[name='score03"+(str(int((i-18)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input03"+(str(int((i-18)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 27<i<33:
                      name_score="[name='score04"+(str(int((i-26)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input04"+(str(int((i-26)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 33<i<37:
                      name_score="[name='score05"+(str(int((i-32)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input05"+(str(int((i-32)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 37<i<41:
                      name_score="[name='score06"+(str(int((i-36)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input06"+(str(int((i-36)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 41<i<45:
                      name_score="[name='score07"+(str(int((i-40)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input07"+(str(int((i-40)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 45<i<47:
                      name_score="[name='score08"+(str(int((i-44)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input08"+(str(int((i-44)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 47<i<51:
                      name_score="[name='score09"+(str(int((i-46)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input09"+(str(int((i-46)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 51<i<53:
                      name_score="[name='score10"+(str(int((i-50)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input10"+(str(int((i-50)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 53<i<71:
                      name_score="[name='score11"+(str(int((i-52)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input11"+(str(int((i-52)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                  elif typ_jd=='候车（机、船）室':
                    if 0<i<9:
                      name_score="[name='score01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input01"+(str(int(i/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 9<i<19:
                      name_score="[name='score02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input02"+(str(int((i-8)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 19<i<27:
                      name_score="[name='score03"+(str(int((i-18)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input03"+(str(int((i-18)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 27<i<33:
                      name_score="[name='score04"+(str(int((i-26)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input04"+(str(int((i-26)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 33<i<37:
                      name_score="[name='score04"+(str(int((i-32)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input04"+(str(int((i-32)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 37<i<41:
                      name_score="[name='score05"+(str(int((i-36)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input05"+(str(int((i-36)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 41<i<45:
                      name_score="[name='score06"+(str(int((i-40)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input06"+(str(int((i-40)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 45<i<51:
                      name_score="[name='score07"+(str(int((i-44)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input07"+(str(int((i-44)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 51<i<53:
                      name_score="[name='score08"+(str(int((i-50)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input08"+(str(int((i-50)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    elif 53<i<77:
                      name_score="[name='score09"+(str(int((i-52)/2)).rjust(2,'0'))+"']"
                      value_score="[value='"+trueorfalse(i,list_score[i])+"']"
                      input_score="[name='input09"+(str(int((i-52)/2)).rjust(2,'0'))+"']"
                      result(score,i)
                    
                el_sum=browser.find_element_by_xpath("//a[contains(text(),'计算结果')]")
                el_sum.click()
                #填入监督日期
                el_supdate=browser.find_element_by_name('supScores.supDate')
                browser.execute_script('arguments[0].removeAttribute(\"readonly\")',el_supdate)
                el_supdate.clear()
                #el_supdate.send_keys(list_score[152][0:4]+'-'+list_score[152][5:7]+'-'+list_score[152][8:10]+' '+) #输入监督日期（当前时间往前推200天）
                if typ_jd=='餐饮服务':
                  el_supdate.send_keys(list_score[153])
                if typ_jd=='食品生产':
                  el_supdate.send_keys(list_score[119])
                elif typ_jd=='饮用水供应':
                  el_supdate.send_keys(list_score[79])
                elif typ_jd=='食品销售':
                  el_supdate.send_keys(list_score[71])
                elif typ_jd=='住宿业':
                  el_supdate.send_keys(list_score[75])
                elif typ_jd=='候车（机、船）室':
                  el_supdate.send_keys(list_score[81])
                el_pfjgclick=browser.find_element_by_xpath("//label[contains(text(),'评分结果')]")
                el_pfjgclick.click()
#记录评分结果并将近12个月的平均分算好
                el_save=browser.find_element_by_xpath("//button[contains(text(),'保存')]")
                el_save.click()
                el_sumbit_2=browser.find_element_by_xpath("//a[contains(text(),'确定')]")
                el_sumbit_2.click()'''
                print(list_score[1].ljust(20,'…') + '已完成录入')

                #录完一家做记录并向后填写‘-’
                shop_row = 3
                while schedul_sht.cell(shop_row,1).value != list_score[1]:
                    shop_row += 1
                schedul_sht.cell(shop_row,(mm_2019 * 3)).value = '√'
                if schedul_sht.cell(shop_row,2).value == 'A级':
                    for i in range(1,6):
                        schedul_sht.cell(shop_row,((mm_2019 + i) * 3)).value = '-'
                elif schedul_sht.cell(shop_row,2).value == 'B级':
                    for i in range(1,3):
                        schedul_sht.cell(shop_row,((mm_2019 + i) * 3)).value = '-'
                elif schedul_sht.cell(shop_row,2).value == '未定级':
                    for i in range(1,2):
                        schedul_sht.cell(shop_row,((mm_2019 + i) * 3)).value = '-'
                schedul_xlsx.save(path_schedul)

            #把该文件放入下一层文件夹中
            shutil.move(file_name,dir_save)
            print('已将文件' + file_name + '移入' + dir_save)
            print('======================================================')


#遍历进度表，如果有单位监督记录为空，先根据等级向前找√记录，再填-，如果还是空，则提示
for mm_row in range(3,schedul_sht.max_row+1):
    mm_col = mm_2019
    while (schedul_sht.cell(mm_row,(mm_col * 3 + 1)).value == None) or (schedul_sht.cell(mm_row,(mm_col * 3 + 1)).value == '-'):
        mm_col -= 1
        if mm_col == 0:
            print(path_schedul + '中未发现' + schedul_sht.cell(mm_row,1).value + '的卫生监督记录，请填写至少一次，否则无法纳入统计！')
            mm_row += 1
            break
    if schedul_sht.cell(mm_row,2).value == 'A级':
        for i in range(1,6):
            schedul_sht.cell(mm_row,((mm_col + i) * 3 + 1)).value = '-'
    elif schedul_sht.cell(mm_row,2).value == 'B级':
        for i in range(1,3):
            schedul_sht.cell(mm_row,((mm_col + i) * 3 + 1)).value = '-'
    elif schedul_sht.cell(mm_row,2).value == '未定级':
        for i in range(1,2):
            schedul_sht.cell(mm_row,((mm_col + i) * 3 + 1)).value = '-'
schedul_xlsx.save(path_schedul)
print('======================================================')


#读取进度表，显示提示
shop_todo = []
shop_nottodo = []
shop_finish = []
for mm_row in range(3,schedul_sht.max_row + 1):
    if schedul_sht.cell(mm_row,(mm_2019 * 3 + 1)).value == '√':
        shop_finish.append(schedul_sht.cell(mm_row,1).value)
    elif schedul_sht.cell(mm_row,(mm_2019 * 3 + 1)).value == '-':
        shop_nottodo.append(schedul_sht.cell(mm_row,1).value)
    elif schedul_sht.cell(mm_row,(mm_2019 * 3 + 1)).value == None:
        shop_todo.append(schedul_sht.cell(mm_row,1).value)
print(yyyy + '年' + mm + '月' + '卫生监督情况如下：')
print('不必监管：')
for i in shop_nottodo:
    print(i)
print('===================')
print('本月已完成：')
for i in shop_finish:
    print(i)
print('===================')
print('本月需监管：')
for i in shop_todo:
    print(i)
print('======================================================')


#编写总结
for root,dirs,files in os.walk(dir_save):
    for file in files:
        files_save.append(os.path.join(root,file))

shop_name = []
shop_coun = 0
wenti_coun = 0
employee_coun = 0
for file in files_save:
    wb_save = openpyxl.load_workbook(file)
    wb_save_sht = wb_save['Sheet1']
    for c in range(1,wb_save_sht.max_column + 1):
        if wb_save_sht.cell(1,c).value == '员工数':
            yg_c = c
    for r in range(2,wb_save_sht.max_row + 1):
        if wb_save_sht.cell(r,2).value in shop_name:
            print('监测到重复的单位名称，统计结果可能不准确，请确认！')
            print('重复单位名称为：' + wb_save_sht.cell(r,2).value)
        shop_name.append(wb_save_sht.cell(r,2).value)
        shop_coun += 1
        employee_coun += wb_save_sht.cell(r,yg_c).value
        for c in range (1,wb_save_sht.max_column + 1):
            if wb_save_sht.cell(r,c).value == '不符合':
                wenti_coun += 1
    wb_save.save(file)


shop_coun = str(shop_coun)
wenti_coun = str(wenti_coun)
employee_coun = str(employee_coun)
txt = '本月总结：\n\
企业共X家，\n\
开展卫生监督'+ shop_coun + '次，\n\
监管' + employee_coun + '人，\n\
快速检测X次X个项目，\n\
发现阳性问题' + wenti_coun +'个。\n\
采样送检X批次。\n\n\
开展开展鼠类夹夜法、鼠笼法及蚤类、寄生蜱、螨类监测一次、\n\
蚊类二氧化碳诱蚊灯监测两次、\n\
蚊类诱卵器监测一次、\n\
蠓类紫外灯监测一次，\n\
捕获均为0。\n\n\
对西宁机场T1、T2航站楼及贵宾厅开展公共场所空气质量监测，\n\
共对X个点位进行X项监测，\n\
发现不合格X项\n\n\n'
txt_write = open(os.path.join(path_sum,(yyyymm + '汇总.txt')),'a',encoding = 'utf-8')
txt_write.write(txt)
txt_write.close()
print(txt)
print('已生成文件' + os.path.join(path_sum,(yyyymm + '汇总.txt')))


#计算当月是否有需要调整评级的单位并给出建议


input('按回车键退出')
